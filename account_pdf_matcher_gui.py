import sys
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QLabel, QPushButton, QFileDialog, QLineEdit, QTableWidget, QTableWidgetItem, QMessageBox, QProgressBar, QComboBox
)
from PyQt5.QtCore import Qt
import os
import re
import pandas as pd
from openpyxl import load_workbook
import fitz  # PyMuPDF
from datetime import datetime

class AccountPDFMatcherApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Account-PDF Matcher')
        self.setGeometry(100, 100, 800, 600)
        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        # File selection
        self.excel_label = QLabel('No Excel file selected')
        self.excel_btn = QPushButton('Select Excel File')
        self.excel_btn.clicked.connect(self.select_excel)
        self.layout.addWidget(self.excel_label)
        self.layout.addWidget(self.excel_btn)

        self.pdf_label = QLabel('No PDF files selected')
        self.pdf_btn = QPushButton('Select PDF Files')
        self.pdf_btn.clicked.connect(self.select_pdfs)
        self.layout.addWidget(self.pdf_label)
        self.layout.addWidget(self.pdf_btn)

        # Sheet selection
        self.sheet_label = QLabel('Select Sheet:')
        self.sheet_combo = QComboBox()
        self.layout.addWidget(self.sheet_label)
        self.layout.addWidget(self.sheet_combo)
        self.sheet_label.hide()
        self.sheet_combo.hide()

        # User input for account number columns and starting row
        self.account_col_range_label = QLabel('Account Number Columns (e.g., B-D):')
        self.account_col_range_input = QLineEdit()
        self.account_col_range_input.setPlaceholderText('B-D')
        self.layout.addWidget(self.account_col_range_label)
        self.layout.addWidget(self.account_col_range_input)

        self.start_row_label = QLabel('Starting Row (e.g., 8):')
        self.start_row_input = QLineEdit()
        self.start_row_input.setPlaceholderText('8')
        self.layout.addWidget(self.start_row_label)
        self.layout.addWidget(self.start_row_input)

        # User input for flag and earnings columns
        self.flag_col_label = QLabel('Flag Column (e.g., H):')
        self.flag_col_input = QLineEdit()
        self.flag_col_input.setPlaceholderText('H')
        self.layout.addWidget(self.flag_col_label)
        self.layout.addWidget(self.flag_col_input)

        self.earnings_col_label = QLabel('Earnings Column (e.g., M):')
        self.earnings_col_input = QLineEdit()
        self.earnings_col_input.setPlaceholderText('M')
        self.layout.addWidget(self.earnings_col_label)
        self.layout.addWidget(self.earnings_col_input)

        # PDF-date mapping table
        self.pdf_date_table = QTableWidget()
        self.pdf_date_table.setColumnCount(3)
        self.pdf_date_table.setHorizontalHeaderLabels(['PDF File', 'Start Date (MM/DD/YYYY)', 'End Date (MM/DD/YYYY)'])
        self.layout.addWidget(QLabel('PDF-Date Mapping:'))
        self.layout.addWidget(self.pdf_date_table)

        # Run button and progress
        self.run_btn = QPushButton('Run Matching')
        self.run_btn.clicked.connect(self.run_matching)
        self.layout.addWidget(self.run_btn)
        self.progress = QProgressBar()
        self.layout.addWidget(self.progress)

        # Internal state
        self.excel_path = None
        self.pdf_paths = []
        self.header = []
        self.sheetnames = []

    def select_excel(self):
        path, _ = QFileDialog.getOpenFileName(self, 'Select Excel File', '', 'Excel Files (*.xlsx *.xls)')
        if path:
            self.excel_path = path
            self.excel_label.setText(f'Selected: {path}')
            # Populate sheet dropdown
            try:
                wb = load_workbook(self.excel_path, read_only=True)
                self.sheetnames = wb.sheetnames
                self.sheet_combo.clear()
                self.sheet_combo.addItems(self.sheetnames)
                self.sheet_label.show()
                self.sheet_combo.show()
            except Exception as e:
                QMessageBox.critical(self, 'Excel Error', f'Failed to read sheets: {e}')
                self.sheet_label.hide()
                self.sheet_combo.hide()

    def select_pdfs(self):
        paths, _ = QFileDialog.getOpenFileNames(self, 'Select PDF Files', '', 'PDF Files (*.pdf)')
        if paths:
            self.pdf_paths = paths
            self.pdf_label.setText(f'Selected: {len(paths)} PDF(s)')
            self.populate_pdf_date_table()

    def populate_pdf_date_table(self):
        self.pdf_date_table.setRowCount(len(self.pdf_paths))
        for i, pdf in enumerate(self.pdf_paths):
            self.pdf_date_table.setItem(i, 0, QTableWidgetItem(pdf))
            # Auto-fill date range based on filename
            base = os.path.basename(pdf)
            m = re.search(r'(6|12)-(\d{4})', base)
            if m:
                period, year = int(m.group(1)), int(m.group(2))
                if period == 6:
                    start_date = f'01/01/{year}'
                    end_date = f'06/30/{year}'
                elif period == 12:
                    start_date = f'07/01/{year}'
                    end_date = f'12/31/{year}'
                else:
                    start_date = ''
                    end_date = ''
            else:
                start_date = ''
                end_date = ''
            self.pdf_date_table.setItem(i, 1, QTableWidgetItem(start_date))
            self.pdf_date_table.setItem(i, 2, QTableWidgetItem(end_date))

    def col_letter_to_index(self, letter):
        letter = letter.upper()
        idx = 0
        for c in letter:
            if 'A' <= c <= 'Z':
                idx = idx * 26 + (ord(c) - ord('A') + 1)
        return idx - 1

    def parse_col_range(self, col_range_str):
        if '-' in col_range_str:
            start, end = col_range_str.split('-')
            start_idx = self.col_letter_to_index(start.strip())
            end_idx = self.col_letter_to_index(end.strip())
            return list(range(start_idx, end_idx + 1))
        else:
            return [self.col_letter_to_index(col_range_str.strip())]

    def run_matching(self):
        if not self.excel_path or not self.pdf_paths:
            QMessageBox.warning(self, 'Missing Files', 'Please select both an Excel file and at least one PDF.')
            return
        col_range_str = self.account_col_range_input.text().strip()
        start_row_str = self.start_row_input.text().strip()
        flag_col_str = self.flag_col_input.text().strip()
        earnings_col_str = self.earnings_col_input.text().strip()
        if not col_range_str or not start_row_str.isdigit() or not flag_col_str or not earnings_col_str:
            QMessageBox.warning(self, 'Input Error', 'Please specify the account number columns, starting row, flag column, and earnings column.')
            return
        account_col_indices = self.parse_col_range(col_range_str)  # 0-based
        start_row = int(start_row_str)
        date_col_idx = 0
        # Use user-inputted columns for flag and earnings
        flag_col_idx = self.col_letter_to_index(flag_col_str)
        earnings_col_idx = self.col_letter_to_index(earnings_col_str)

        # Sheet selection
        if self.sheet_combo.count() == 0:
            QMessageBox.warning(self, 'Sheet Error', 'No sheet selected.')
            return
        sheet_name = self.sheet_combo.currentText()

        # Parse PDF-date mapping from table
        pdf_date_map = []
        for row in range(self.pdf_date_table.rowCount()):
            pdf_path = self.pdf_date_table.item(row, 0).text() if self.pdf_date_table.item(row, 0) else ''
            start_str = self.pdf_date_table.item(row, 1).text() if self.pdf_date_table.item(row, 1) else ''
            end_str = self.pdf_date_table.item(row, 2).text() if self.pdf_date_table.item(row, 2) else ''
            try:
                start_date = datetime.strptime(start_str, '%m/%d/%Y').date() if start_str else None
                end_date = datetime.strptime(end_str, '%m/%d/%Y').date() if end_str else None
            except Exception as e:
                QMessageBox.critical(self, 'Date Error', f'Error parsing date for PDF {os.path.basename(pdf_path)}: {e}')
                return
            if pdf_path and start_date and end_date:
                pdf_date_map.append((pdf_path, start_date, end_date))
        if not pdf_date_map:
            QMessageBox.warning(self, 'PDF-Date Mapping', 'Please fill in the date ranges for each PDF.')
            return

        # Build PDF account→earnings maps
        self.progress.setValue(0)
        pdf_account_earnings = {}
        for i, (pdf_path, _, _) in enumerate(pdf_date_map):
            self.progress.setValue(int(10 + 20 * i / max(1, len(pdf_date_map))))
            try:
                pdf_account_earnings[pdf_path] = self.extract_account_earnings_from_pdf(pdf_path)
            except Exception as e:
                QMessageBox.critical(self, 'PDF Error', f'Error processing PDF {os.path.basename(pdf_path)}: {e}')
                return

        # Open Excel for update
        try:
            wb = load_workbook(self.excel_path)
            if sheet_name not in wb.sheetnames:
                QMessageBox.critical(self, 'Sheet Error', f'Sheet {sheet_name} not found in workbook.')
                return
            ws = wb[sheet_name]
        except Exception as e:
            QMessageBox.critical(self, 'Excel Error', f'Error opening Excel: {e}')
            return

        n_rows = ws.max_row
        n_matched = 0
        for row_idx in range(start_row, n_rows + 1):
            self.progress.setValue(int(30 + 70 * (row_idx-start_row) / max(1, n_rows-start_row)))
            row = [ws.cell(row=row_idx, column=col+1).value for col in range(ws.max_column)]
            date_cell = row[date_col_idx]
            if not date_cell:
                continue
            if isinstance(date_cell, datetime):
                row_date = date_cell.date()
            else:
                try:
                    row_date = pd.to_datetime(date_cell).date()
                except Exception:
                    continue
            pdf_path = None
            for path, start, end in pdf_date_map:
                if start <= row_date <= end:
                    pdf_path = path
                    break
            if not pdf_path:
                ws.cell(row=row_idx, column=flag_col_idx+1, value='M')
                ws.cell(row=row_idx, column=earnings_col_idx+1, value=0.0)
                continue
            found = False
            total_earnings = 0.0
            # Debug: print all normalized Excel account numbers for this row
            excel_accts = []
            for acct_idx in account_col_indices:
                acct_val = row[acct_idx]
                # Convert float to int if possible, then to string
                if isinstance(acct_val, float) and acct_val.is_integer():
                    acct_key = str(int(acct_val))
                elif acct_val is not None:
                    acct_key = str(acct_val).strip()
                else:
                    acct_key = ''
                # Standardize: pad to 5 digits if all digits, else leave as is
                if acct_key.isdigit() and len(acct_key) < 5:
                    acct_key = acct_key.zfill(5)
                excel_accts.append(acct_key)
            print(f"Row {row_idx}: Excel account numbers (normalized): {excel_accts}")
            # Get PDF account earnings dict for this PDF
            pdf_acct_earnings = pdf_account_earnings.get(pdf_path, {})
            print(f"PDF {os.path.basename(pdf_path)}: Extracted accounts: {list(pdf_acct_earnings.keys())}")
            for acct_key in excel_accts:
                # Try direct match
                earnings_list = pdf_acct_earnings.get(acct_key, [])
                # If not found, try matching with/without leading zeros
                if not earnings_list:
                    # Try matching with leading zeros (up to 10 digits)
                    for k in pdf_acct_earnings.keys():
                        if k.lstrip('0') == acct_key.lstrip('0'):
                            earnings_list = pdf_acct_earnings[k]
                            print(f"  Matched Excel {acct_key} to PDF {k}")
                            break
                else:
                    print(f"  Direct match found for {acct_key}")
                if earnings_list:
                    found = True
                    total_earnings += sum(earnings_list)
            if found:
                print(f"  -> Row {row_idx}: MATCHED, total earnings: {total_earnings}")
                ws.cell(row=row_idx, column=flag_col_idx+1, value='x')
                ws.cell(row=row_idx, column=earnings_col_idx+1, value=total_earnings)
                n_matched += 1
            else:
                print(f"  -> Row {row_idx}: NO MATCH")
                ws.cell(row=row_idx, column=flag_col_idx+1, value='M')
                ws.cell(row=row_idx, column=earnings_col_idx+1, value=0.0)

        self.progress.setValue(100)
        save_path, _ = QFileDialog.getSaveFileName(self, 'Save Updated Excel', os.path.splitext(self.excel_path)[0] + '_updated.xlsx', 'Excel Files (*.xlsx)')
        if save_path:
            try:
                wb.save(save_path)
                QMessageBox.information(self, 'Done', f'Updated Excel saved. {n_matched} rows matched.')
            except Exception as e:
                QMessageBox.critical(self, 'Save Error', f'Error saving Excel: {e}')
        else:
            QMessageBox.information(self, 'Cancelled', 'Save cancelled.')

    def extract_account_earnings_from_pdf(self, pdf_path):
        doc = fitz.open(pdf_path)
        full_text = []
        for page in doc:
            full_text += page.get_text("text").splitlines()
        account_earnings = {}
        capture = False
        line_count = len(full_text)
        for i, raw_line in enumerate(full_text):
            line = raw_line.strip()
            if "Royalty Accounts Information" in line:
                capture = True
                continue
            if capture and line.lower().startswith("total royalty earnings"):
                capture = False
                continue
            if not capture:
                continue
            m_acct = re.fullmatch(r"(\d+)", line)
            if m_acct:
                raw_acct = m_acct.group(1)
                # Standardize: pad to 5 digits if all digits and < 5, else leave as is
                if raw_acct.isdigit() and len(raw_acct) < 5:
                    acct = raw_acct.zfill(5)
                else:
                    acct = raw_acct
                # Also store with and without leading zeros for matching
                acct_variants = set([acct, acct.lstrip('0')])
                # look ahead up to 20 lines for the first $xx.xx
                for j in range(i+1, min(i+20, line_count)):
                    nxt = full_text[j].strip()
                    if not nxt:
                        continue
                    if re.fullmatch(r"\d+", nxt) or nxt.lower().startswith("total royalty earnings"):
                        break
                    m_amt = re.search(r"\$([\d,]+\.\d{2})", nxt)
                    if m_amt:
                        amt = float(m_amt.group(1).replace(",", ""))
                        for variant in acct_variants:
                            account_earnings.setdefault(variant, []).append(amt)
                        break
        # Debug: print all extracted accounts and earnings
        print(f"Extracted from {os.path.basename(pdf_path)}: {account_earnings}")
        return account_earnings

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = AccountPDFMatcherApp()
    window.show()
    sys.exit(app.exec_()) 